#!/usr/bin/env python
"""
Direct test of Excel generation core functionality
"""
import pandas as pd
import openpyxl
from io import BytesIO

def test_excel_creation():
    """Test if we can create Excel files with student data"""
    print("Testing Excel creation...")
    
    # Sample data mimicking what our system should generate
    student_data = [
        {'S.N.': 1, 'Roll Number': 'BCE001', 'Student Name': 'John Doe', 'Subject': 'Machine Learning'},
        {'S.N.': 2, 'Roll Number': 'BCE002', 'Student Name': 'Jane Smith', 'Subject': 'Machine Learning'},
        {'S.N.': 3, 'Roll Number': 'BCE003', 'Student Name': 'Bob Johnson', 'Subject': 'Machine Learning'},
    ]
    
    # Create DataFrame
    df = pd.DataFrame(student_data)
    print("Student data:")
    print(df)
    print()
    
    # Create Excel file in memory
    excel_buffer = BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # Main student list
        df.to_excel(writer, sheet_name='Students', index=False)
        
        # Subject info
        subject_info = pd.DataFrame({
            'Info': ['Subject Name', 'Total Students', 'Generated Date'],
            'Value': ['Machine Learning', len(df), '2025-07-20']
        })
        subject_info.to_excel(writer, sheet_name='Subject Info', index=False)
        
        # Attendance template
        attendance_df = df.copy()
        attendance_df['Attendance'] = ''
        attendance_df['Remarks'] = ''
        attendance_df.to_excel(writer, sheet_name='Attendance', index=False)
    
    # Get the Excel data
    excel_data = excel_buffer.getvalue()
    
    # Save test file
    test_filename = '/tmp/test_machine_learning_students.xlsx'
    with open(test_filename, 'wb') as f:
        f.write(excel_data)
    
    print(f"✅ Excel file created: {test_filename}")
    print(f"   File size: {len(excel_data)} bytes")
    
    # Verify the file can be opened
    try:
        wb = openpyxl.load_workbook(test_filename)
        print(f"   Worksheets: {wb.sheetnames}")
        
        # Check student sheet content
        ws = wb['Students']
        print(f"   Students sheet has {ws.max_row-1} students")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"❌ Error verifying Excel file: {e}")
        return False

if __name__ == "__main__":
    success = test_excel_creation()
    print(f"\nTest result: {'PASSED' if success else 'FAILED'}")
